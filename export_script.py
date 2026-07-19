from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill

from app.data.database.database import session_local, Base, engine
from app.data.database.tables.evaluation_db import EvaluationDb
from app.domain.models.evaluation.evaluation import Evaluation
from app.domain.models.evaluation.nutrition_search_info import NutritionSearchInfo
from app.domain.models.recipe.nutrition import Nutrition

def export():
    Base.metadata.create_all(bind=engine)
    db = session_local()
    try:
        evaldb_list = db.query(EvaluationDb).all()

        evaluation_rows = []
        nutrition_rows = []
        search_history_rows = []
        nutrients = list(Nutrition.model_fields.keys())  # Creates a list of all the fields in the Nutrition model
        for item in evaldb_list:
            evaluation = Evaluation.model_validate(item.evaluation)
            valid = (
                    len(evaluation.original_search_info.failed_ingredients) == 0
                    and len(evaluation.generated_search_info.failed_ingredients) == 0
            )

            evaluation_rows.append({
                "uuid": item.uuid,
                "url": item.url,
                "cosine similarity": evaluation.cosine_similarity,
                "ingredient overlap": evaluation.ingredient_overlap,
                "original lookup failure (%)": _calculation_failure_rate(evaluation.original_search_info),
                "original failed ingredients": evaluation.original_search_info.failed_ingredients,
                "original skipped ingredients": evaluation.original_search_info.skipped_ingredients,
                "generated lookup failure (%)": _calculation_failure_rate(evaluation.generated_search_info),
                "generated failed ingredients": evaluation.generated_search_info.failed_ingredients,
                "generated skipped ingredients": evaluation.generated_search_info.skipped_ingredients,
                "valid": valid
            })

            for search in evaluation.original_search_info.matched_ingredients:
                search_history_rows.append({
                    "uuid": item.uuid,
                    "recipe type": "original",
                    "query": search.search_query,
                    "result": search.result_description
                })
            for search in evaluation.generated_search_info.matched_ingredients:
                search_history_rows.append({
                    "uuid": item.uuid,
                    "recipe type": "generated",
                    "query": search.search_query,
                    "result": search.result_description
                })


            for nutrient in nutrients:
                nutrition_rows.append({
                    "uuid": item.uuid,
                    "nutrient": nutrient,

                    "original recipe": getattr(evaluation.original_web_nutrition, nutrient, None),
                    "generated recipe": getattr(evaluation.generated_ai_nutrition, nutrient, None),

                    "original calculated": getattr(evaluation.original_calculated_nutrition, nutrient, None),
                    "generated calculated": getattr(evaluation.generated_calculated_nutrition, nutrient, None),

                    "recipe change (%)": getattr(evaluation.ai_nutrition_changes, nutrient, None),
                    "calculated change (%)": getattr(evaluation.calculated_nutrition_changes,nutrient, None),
                    "Error (pp)": getattr(evaluation.percentage_point_error,nutrient, None),
                })



        evaluation_sheet = pd.DataFrame(evaluation_rows)
        nutrition_comparison_sheet = pd.DataFrame(nutrition_rows)
        search_history_sheet = pd.DataFrame(search_history_rows)

        output_file = Path("evaluation_export.xlsx")
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            evaluation_sheet.to_excel(writer, sheet_name="Evaluations", index=False)
            nutrition_comparison_sheet.to_excel(writer, sheet_name="Nutrition Comparison", index=False)
            search_history_sheet.to_excel(writer, sheet_name="Search History", index=False)

            #Color invalid lines yellow
            workbook = writer.book
            evaluations_ws = workbook["Evaluations"]
            headers = [cell.value for cell in evaluations_ws[1]]
            valid_col = headers.index("valid")
            for row in evaluations_ws.iter_rows():
                if not row[valid_col].value:
                    for cell in row: cell.fill = PatternFill(fill_type="solid",start_color="FFFF00",end_color="FFFF00")




    finally:
        db.close()

def _calculation_failure_rate(search_info: NutritionSearchInfo) -> float:
    total_ingredients = len(search_info.matched_ingredients) + len(search_info.failed_ingredients)
    failed_lookups = len(search_info.failed_ingredients)
    if total_ingredients == 0:return 1
    return round(failed_lookups / total_ingredients * 100, 1)

if __name__ == "__main__":
    export()